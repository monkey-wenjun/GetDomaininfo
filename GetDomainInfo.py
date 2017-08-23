#!/usr/bin/env python
#-*-coding:utf-8-*-

import requests
import json
import time
from openpyxl import Workbook

# 获取到备案号调用checkarea 判断归属哪个区,N 代表华北,E 代表华东，S 代表华南
def checkarea(area):

    N = ["京", "津", "冀", "晋", "蒙", "辽", "吉", "黑", "陕", "甘", "青", "宁", "新"]
    E = ["沪", "苏", "浙", "皖", "赣", "鲁", "豫", "闽"]
    S = ["粤", "桂", "琼", "鄂", "湘", "渝", "蜀", "黔", "滇"]
    # print(area)
    if area in N:
        return "N"
    if area in E:
        return "E"
    if area in S:
        return "S"

def CheckDomain():

    wb = Workbook()
    ws = wb.active
    # 表名称
    ws.title = "数据"
    # 对应列的名称
    ws['A1'] = "域名"
    ws["B1"] = "企业/个人"
    ws["C1"] = "名称"
    ws["D1"] = "主页地址"
    ws["E1"] = "网站名称"
    ws["F1"] = "备案号"
    ws["G1"] = "地区"

    # 读取域名列表
    with  open("domain") as f:
        while 1:
            lines = f.readlines(100000) # 使用readlines读取数据效率更高
            if not lines:
                break
            for line in lines:
                #接口地址
                api = "http://www.sojson.com/api/beian/"
                # 伪造个user-agent 避免被接口判断为机器人请求，并且加上no-cache的头，避免读取到缓存缓存
                headers = {"User-Agent":'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.90 Safari/537.36','cache-control':'no-cache'}
               #调用接口获取备案信息
                try:
                    r = requests.get(api+line.strip('\n'),headers=headers)
                except (requests.ConnectionError, IndexError, UnicodeEncodeError,TimeoutError):
                    print("请求异常，无法连接服务器")
                except requests.HTTPError as f:
                    print('请求异常，HTTP错误')
                finally:
                    info = r.json()
                    # 如果备案type值为200，说明有备案信息，获取对应的备案号，公司名等信息，如果不是200 则说明没有获取到备案信息
                    if info['type'] == 200:
                        print(info)
                        # 获取备案主体
                        getnature = info['nature']
                        # 获取名称
                        getdomain = info['domain'].strip()
                        print(getdomain)
                        #获取主体名字名字

                        if getnature =="个人":
                            getname = "个人"
                        elif getnature =="企业":
                            getname = info['name']
                        else:
                            getname = info['name']

                        # 获取备案号
                        getnoicp = info['nowIcp']
                        getarea = checkarea(info['nowIcp'][0])
                        # 获取主页地址
                        getindexurl = info['indexUrl']
                        # 获取网站名称
                        getsitename = info['sitename']
                        # 根据备案号的第一个下标的值来判断所属的大区是华北、华东还是华南，然后写入到表格的没一行中
                        if getarea == "N":
                            N = "华北"
                            ws.append([getdomain,getnature,getname, getindexurl,getsitename,getnoicp,N])
                        if getarea =="E":
                            E = "华东"
                            ws.append([getdomain, getnature, getname, getindexurl, getsitename, getnoicp,E])
                        if getarea == "S":
                            S = "华南"
                            ws.append([getdomain,getnature,getname,getindexurl,getsitename,getnoicp,S])
                    else:
                        ws.append([line,'无备案信息' ,'' ,'','' ])
                    # 保存表格为xlsx
                    wb.save("客户信息1-1.xlsx")
                    # 避免被接口拒绝，3秒请求一次
                    time.sleep(3)
                    # 关闭表格的数据写入
                    wb.close()
    print("Job Done!")

CheckDomain()
